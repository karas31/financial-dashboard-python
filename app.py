import streamlit as st
import yfinance as yf
import pandas as pd
import plotly.graph_objects as go
import json
import os
from io import BytesIO
from streamlit_autorefresh import st_autorefresh
from yfinance.exceptions import YFRateLimitError
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# -----------------------------
# Page Setup
# -----------------------------
st.set_page_config(page_title="Financial Dashboard", layout="wide")
st_autorefresh(interval=300000, key="refresh")

PINS_FILE = "pins.json"
DEFAULT_PINS = ["AAPL", "MSFT", "NVDA"]


# -----------------------------
# Helper Functions
# -----------------------------
def load_pins():
    if os.path.exists(PINS_FILE):
        try:
            with open(PINS_FILE, "r") as file:
                pins = json.load(file)

            while len(pins) < 3:
                pins.append(DEFAULT_PINS[len(pins)])

            return pins[:3]
        except Exception:
            return DEFAULT_PINS

    return DEFAULT_PINS


def save_pins(pins):
    with open(PINS_FILE, "w") as file:
        json.dump(pins, file)


def clean_columns(df):
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    return df


def format_market_cap(value):
    if not value:
        return "N/A"

    if value >= 1_000_000_000_000:
        return f"${value / 1_000_000_000_000:.2f}T"

    if value >= 1_000_000_000:
        return f"${value / 1_000_000_000:.2f}B"

    if value >= 1_000_000:
        return f"${value / 1_000_000:.2f}M"

    return f"${value:,.2f}"


@st.cache_data(ttl=900)
def get_price_data(symbol, period="1y"):
    try:
        df = yf.download(symbol, period=period,
                         auto_adjust=False, progress=False)
        return clean_columns(df)
    except YFRateLimitError:
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600)
def build_stock_summary(symbol):
    stock_data = get_price_data(symbol, period="1y")

    if stock_data.empty:
        return None

    stock_data["Daily Return"] = stock_data["Close"].pct_change()

    current_price = float(stock_data["Close"].iloc[-1])
    first_price = float(stock_data["Close"].iloc[0])
    stock_return = ((current_price - first_price) / first_price) * 100
    stock_volatility = float(stock_data["Daily Return"].std()) * 100

    return {
        "Ticker": symbol,
        "Current Price": round(current_price, 2),
        "1Y Return %": round(stock_return, 2),
        "Volatility %": round(stock_volatility, 2),
    }
    return {
        "Ticker": symbol,
        "Current Price": round(current_price, 2),
        "1Y Return %": round(stock_return, 2),
        "Volatility %": round(stock_volatility, 2),
    }


@st.cache_data(ttl=900)
def get_price_data(symbol, period="1y"):
    try:
        df = yf.download(symbol, period=period,
                         auto_adjust=False, progress=False)
        return clean_columns(df)
    except YFRateLimitError:
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600)
def get_stock_info(symbol):
    try:
        return yf.Ticker(symbol).info
    except YFRateLimitError:
        return {}
    except Exception:
        return {}


@st.cache_data(ttl=3600)
def get_stock_news(symbol):
    try:
        return yf.Ticker(symbol).news
    except YFRateLimitError:
        return []
    except Exception:
        return []


def build_stock_summary(symbol):
    stock_data = get_price_data(symbol, period="1y")

    if stock_data.empty:
        return None

    stock_data["Daily Return"] = stock_data["Close"].pct_change()

    current_price = float(stock_data["Close"].iloc[-1])
    first_price = float(stock_data["Close"].iloc[0])
    stock_return = ((current_price - first_price) / first_price) * 100
    stock_volatility = float(stock_data["Daily Return"].std()) * 100

    return {
        "Ticker": symbol,
        "Current Price": round(current_price, 2),
        "1Y Return %": round(stock_return, 2),
        "Volatility %": round(stock_volatility, 2),
    }


# -----------------------------
# Sidebar / Watchlist
# -----------------------------
saved_pins = load_pins()

st.title("Financial Dashboard")
st.sidebar.header("Pinned Watchlist")

pin1 = st.sidebar.text_input("Pinned Stock 1", saved_pins[0]).strip().upper()
pin2 = st.sidebar.text_input("Pinned Stock 2", saved_pins[1]).strip().upper()
pin3 = st.sidebar.text_input("Pinned Stock 3", saved_pins[2]).strip().upper()

pinned_stocks = [pin for pin in [pin1, pin2, pin3] if pin]

if not pinned_stocks:
    pinned_stocks = ["AAPL"]

save_pins([pin1, pin2, pin3])

default_stocks = ["GOOG", "TSLA", "AMZN", "META"]
stock_options = list(dict.fromkeys(pinned_stocks + default_stocks))

if "selected_ticker" not in st.session_state:
    st.session_state["selected_ticker"] = pinned_stocks[0]

if st.session_state["selected_ticker"] not in stock_options:
    st.session_state["selected_ticker"] = stock_options[0]


# -----------------------------
# Quick Access
# -----------------------------
st.subheader("Quick Access")

quick_cols = st.columns(len(pinned_stocks))

for i, symbol in enumerate(pinned_stocks):
    if quick_cols[i].button(symbol):
        st.session_state["selected_ticker"] = symbol

ticker = st.selectbox(
    "Choose a stock ticker:",
    stock_options,
    index=stock_options.index(st.session_state["selected_ticker"]),
)

st.session_state["selected_ticker"] = ticker


# -----------------------------
# Main Stock Data
# -----------------------------
data = get_price_data(ticker, period="1y")

if data.empty:
    st.error(f"No data found for ticker: {ticker}")
    st.stop()

info = get_stock_info(ticker)

if not info:
    st.warning(
        "Some company information is temporarily unavailable because Yahoo Finance is rate-limiting requests. "
        "Price data may still work. Please try again later."
    )

company_name = info.get("longName", ticker)
sector = info.get("sector", "N/A")
industry = info.get("industry", "N/A")
market_cap = info.get("marketCap")
summary = info.get("longBusinessSummary", "No company summary available.")

pe_ratio = info.get("trailingPE")
forward_pe = info.get("forwardPE")
beta = info.get("beta")
dividend_yield = info.get("dividendYield")
fifty_two_week_high = info.get("fiftyTwoWeekHigh")
fifty_two_week_low = info.get("fiftyTwoWeekLow")

data["Daily Return"] = data["Close"].pct_change()
data["50 Day MA"] = data["Close"].rolling(window=50).mean()
data["200 Day MA"] = data["Close"].rolling(window=200).mean()

latest_price = float(data["Close"].iloc[-1])
start_price = float(data["Close"].iloc[0])
total_return = (latest_price - start_price) / start_price
volatility = float(data["Daily Return"].std())

if volatility < 0.015:
    risk_level = "🟢 Low Risk"
elif volatility < 0.025:
    risk_level = "🟡 Medium Risk"
else:
    risk_level = "🔴 High Risk"


# -----------------------------
# Main Header Metrics
# -----------------------------
st.header(f"{company_name} ({ticker})")

metric_cols = st.columns(5)

metric_cols[0].metric("Current Price", f"${latest_price:.2f}")
metric_cols[1].metric("1-Year Return", f"{total_return:.2%}")
metric_cols[2].metric("Daily Volatility", f"{volatility:.2%}")
metric_cols[3].metric("Risk Level", risk_level)
metric_cols[4].metric("Market Cap", format_market_cap(market_cap))


# -----------------------------
# Tabs
# -----------------------------
overview_tab, portfolio_tab, compare_tab, news_tab, export_tab = st.tabs([
    "Overview",
    "Portfolio",
    "Compare",
    "News",
    "Export"
])


# -----------------------------
# Overview Tab
# -----------------------------
with overview_tab:
    st.subheader("Company Information")

    info_col1, info_col2 = st.columns(2)

    info_col1.write(f"**Sector:** {sector}")
    info_col1.write(f"**Industry:** {industry}")
    info_col2.write(f"**Ticker:** {ticker}")
    info_col2.write("**Refresh Rate:** 30 seconds")

    with st.expander("Company Summary"):
        st.write(summary)

    st.subheader("Valuation & Fundamentals")

    fund_cols = st.columns(6)

    fund_cols[0].metric("P/E Ratio", f"{pe_ratio:.2f}" if pe_ratio else "N/A")
    fund_cols[1].metric(
        "Forward P/E", f"{forward_pe:.2f}" if forward_pe else "N/A")
    fund_cols[2].metric("Beta", f"{beta:.2f}" if beta else "N/A")
    fund_cols[3].metric(
        "52W High", f"${fifty_two_week_high:.2f}" if fifty_two_week_high else "N/A")
    fund_cols[4].metric(
        "52W Low", f"${fifty_two_week_low:.2f}" if fifty_two_week_low else "N/A")
    fund_cols[5].metric(
        "Dividend Yield", f"{dividend_yield * 100:.2f}%" if dividend_yield else "N/A")

    st.subheader("Interactive Price History")

    fig = go.Figure()

    fig.add_trace(
        go.Scatter(
            x=data.index,
            y=data["Close"],
            mode="lines",
            name="Close Price",
        )
    )

    fig.add_trace(
        go.Scatter(
            x=data.index,
            y=data["50 Day MA"],
            mode="lines",
            name="50 Day Moving Average",
        )
    )

    fig.add_trace(
        go.Scatter(
            x=data.index,
            y=data["200 Day MA"],
            mode="lines",
            name="200 Day Moving Average",
        )
    )

    fig.update_layout(
        title=f"{ticker} Price History",
        xaxis_title="Date",
        yaxis_title="Price",
        hovermode="x unified",
        height=500,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
        ),
    )

    st.plotly_chart(fig, use_container_width=True)

    st.subheader("Recent Stock Data")
    st.dataframe(data.tail(), use_container_width=True)

    st.subheader("Analyst Summary")

    if total_return > 0:
        st.success(
            f"{ticker} has generated a positive return of {total_return:.2%} over the last year.")
    else:
        st.warning(
            f"{ticker} has generated a negative return of {total_return:.2%} over the last year.")

    if data["50 Day MA"].iloc[-1] > data["200 Day MA"].iloc[-1]:
        st.info("The stock is currently in an upward trend because the 50-Day Moving Average is above the 200-Day Moving Average.")
    else:
        st.info("The stock is currently in a weaker trend because the 50-Day Moving Average is below the 200-Day Moving Average.")

# -----------------------------
# Portfolio Tab
# -----------------------------
    # -----------------------------
# Portfolio Tab
# -----------------------------
with portfolio_tab:
    st.subheader("Portfolio Tracker")

    portfolio_input_col1, portfolio_input_col2 = st.columns(2)

    shares_owned = portfolio_input_col1.number_input(
        "Shares Owned",
        min_value=0.0,
        value=0.0,
        step=1.0,
    )

    cost_basis = portfolio_input_col2.number_input(
        "Average Cost Basis Per Share",
        min_value=0.0,
        value=0.0,
        step=1.0,
    )

    current_value = shares_owned * latest_price
    total_cost = shares_owned * cost_basis
    gain_loss = current_value - total_cost
    gain_loss_percent = gain_loss / total_cost if total_cost > 0 else 0

    portfolio_cols = st.columns(4)

    portfolio_cols[0].metric("Shares Owned", f"{shares_owned:.2f}")
    portfolio_cols[1].metric("Current Value", f"${current_value:,.2f}")
    portfolio_cols[2].metric("Total Gain/Loss", f"${gain_loss:,.2f}")
    portfolio_cols[3].metric("Gain/Loss %", f"{gain_loss_percent:.2%}")


# -----------------------------
# Compare Tab
# -----------------------------
with compare_tab:
    st.subheader("Stock Comparison Tool")

    comparison_stocks = st.multiselect(
        "Select stocks to compare:",
        stock_options,
        default=pinned_stocks,
    )

    comparison_data = []

    for symbol in comparison_stocks:
        summary_row = build_stock_summary(symbol)

        if summary_row:
            comparison_data.append(summary_row)

    comparison_df = pd.DataFrame(comparison_data)
    st.dataframe(comparison_df, use_container_width=True)

    st.subheader("Pinned Stock Watchlist")

    watchlist_data = []

    for symbol in pinned_stocks:
        summary_row = build_stock_summary(symbol)

        if summary_row:
            watchlist_data.append(summary_row)

    watchlist_df = pd.DataFrame(watchlist_data)
    st.dataframe(watchlist_df, use_container_width=True)


# -----------------------------
# News Tab
# -----------------------------
with news_tab:
    st.subheader("Latest News")

    try:
        news = get_stock_news(ticker)

        if news:
            for article in news[:5]:
                content = article.get("content", {})

                title = content.get("title", "No Title")
                news_summary = content.get("summary", "")
                publisher = content.get("provider", {}).get(
                    "displayName", "Unknown Publisher")

                url = (
                    content.get("clickThroughUrl", {}).get("url")
                    or content.get("canonicalUrl", {}).get("url")
                    or "#"
                )

                st.markdown(f"### [{title}]({url})")
                st.caption(f"Source: {publisher}")

                if news_summary:
                    st.write(news_summary)

                st.divider()

        else:
            st.write("No recent news available.")

    except Exception:
        st.write("Unable to retrieve news.")


# -----------------------------
# Export Tab
# -----------------------------
with export_tab:
    st.subheader("Export Professional Report")

    report_generated = pd.Timestamp.now().strftime("%Y-%m-%d %I:%M %p")

    overview_df = pd.DataFrame({
        "Field": [
            "Report Generated",
            "Company",
            "Ticker",
            "Sector",
            "Industry",
            "Market Cap",
            "Current Price",
            "1-Year Return",
            "Daily Volatility",
            "Risk Level",
            "P/E Ratio",
            "Forward P/E",
            "Beta",
            "52 Week High",
            "52 Week Low",
            "Dividend Yield",
        ],
        "Value": [
            report_generated,
            company_name,
            ticker,
            sector,
            industry,
            format_market_cap(market_cap),
            f"${latest_price:.2f}",
            f"{total_return:.2%}",
            f"{volatility:.2%}",
            risk_level,
            f"{pe_ratio:.2f}" if pe_ratio else "N/A",
            f"{forward_pe:.2f}" if forward_pe else "N/A",
            f"{beta:.2f}" if beta else "N/A",
            f"${fifty_two_week_high:.2f}" if fifty_two_week_high else "N/A",
            f"${fifty_two_week_low:.2f}" if fifty_two_week_low else "N/A",
            f"{dividend_yield * 100:.2f}%" if dividend_yield else "N/A",
        ],
    })

    portfolio_df = pd.DataFrame({
        "Metric": [
            "Shares Owned",
            "Average Cost Basis",
            "Current Value",
            "Total Cost",
            "Gain/Loss",
            "Gain/Loss %",
        ],
        "Value": [
            shares_owned,
            cost_basis,
            current_value,
            total_cost,
            gain_loss,
            gain_loss_percent,
        ],
    })

    price_history_df = data.tail(60).reset_index()

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Overview",
                             index=False, startrow=3)
        portfolio_df.to_excel(
            writer, sheet_name="Portfolio", index=False, startrow=3)
        comparison_df.to_excel(
            writer, sheet_name="Comparison", index=False, startrow=3)
        watchlist_df.to_excel(
            writer, sheet_name="Watchlist", index=False, startrow=3)
        price_history_df.to_excel(
            writer, sheet_name="Price History", index=False, startrow=3)

        sheet_titles = {
            "Overview": f"{company_name} ({ticker}) - Overview",
            "Portfolio": f"{company_name} ({ticker}) - Portfolio Tracker",
            "Comparison": "Stock Comparison Report",
            "Watchlist": "Pinned Stock Watchlist",
            "Price History": f"{company_name} ({ticker}) - Recent Price History",
        }

        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=14, bold=True, color="1F4E78")
        subtitle_font = Font(size=10, italic=True, color="666666")
        thin_border = Border(bottom=Side(style="thin", color="D9EAF7"))

        for sheet_name, worksheet in writer.sheets.items():
            worksheet["A1"] = sheet_titles.get(sheet_name, sheet_name)
            worksheet["A1"].font = title_font

            worksheet["A2"] = f"Generated: {report_generated}"
            worksheet["A2"].font = subtitle_font

            worksheet.freeze_panes = "A5"

            for cell in worksheet[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass

                worksheet.column_dimensions[column_letter].width = min(
                    max_length + 2, 35)

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top")

    st.download_button(
        label="📊 Download Professional Excel Report",
        data=buffer.getvalue(),
        file_name=f"{ticker}_professional_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
